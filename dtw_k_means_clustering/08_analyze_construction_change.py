#!/usr/bin/env python3
"""Analyze segment-level DTW cluster changes before and after tram road controls."""

from __future__ import annotations

import csv
import gzip
import hashlib
import statistics
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

HERE = Path(__file__).resolve().parent
LOCAL_ML_DEPS = HERE / ".deps_ml"
if LOCAL_ML_DEPS.exists():
    sys.path.insert(0, str(LOCAL_ML_DEPS))

import openpyxl


ROOT = HERE.parent
TRAM_SOURCE_DIR = ROOT / "raw_data" / "대전 도시철도 2호선 트램"
CONSTRUCTION_XLSX = TRAM_SOURCE_DIR / "트램_공구별_통제현황.xlsx"
WINDOWS_PATH = HERE / "outputs" / "04_daily_96_windows" / "daily_speed_index_windows.csv.gz"
UNITS_PATH = HERE / "outputs" / "03_speed_index_15min" / "spatial_units_89.csv"
LABELS_PATH = HERE / "outputs" / "07_final_k_selection" / "dtw_cluster_labels.csv"
OUTPUT_DIR = HERE / "outputs" / "08_construction_change_analysis"

EXPECTED_EVENTS = 25
EXPECTED_SAMPLES = 36_154
EXPECTED_UNITS = 89
EXPECTED_BINS = 96
MAX_WINDOW_DAYS = 90
MIN_PERIOD_SAMPLES = 15
RAW_LOW_SHARE_DELTA_THRESHOLD = 0.15
RAW_SPEED_DELTA_THRESHOLD = -0.05
DID_LOW_SHARE_DELTA_THRESHOLD = 0.10
DID_SPEED_DELTA_THRESHOLD = -0.03
LOW_SPEED_LABELS = {"상시 저속형", "출퇴근 집중형"}
SEVERE_LABEL = "상시 저속형"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date: {value!r}")


def format_value(value: Any) -> Any:
    if isinstance(value, float):
        return f"{value:.9f}"
    if isinstance(value, date):
        return value.isoformat()
    return value


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise RuntimeError(f"Refusing to write empty output: {path}")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        for row in rows:
            writer.writerow({key: format_value(value) for key, value in row.items()})


def load_source_events() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(CONSTRUCTION_XLSX, read_only=True, data_only=True)
    control_sheet = workbook["공구별 통제현황"]
    events: list[dict[str, Any]] = []
    source_warnings = {
        17: "7공구 단계별 위치·시작일은 현재 공식 공구 요약과 표기 차이가 있어 원천 XLSX 값을 유지하고 LOW 신뢰도로 분석",
        18: "7공구 단계별 위치·시작일은 현재 공식 공구 요약과 표기 차이가 있어 원천 XLSX 값을 유지하고 LOW 신뢰도로 분석",
        26: "13공구 확장 구간의 실제 확장 시작일이 원천 비고에 미상으로 기재됨; 최초 죽전~신안 구간만 매핑",
    }
    for row_number in range(5, 30):
        values = [control_sheet.cell(row_number, column).value for column in range(1, 11)]
        if not values[0]:
            continue
        events.append(
            {
                "event_id": f"CTRL_{row_number:03d}",
                "source_sheet": "공구별 통제현황",
                "source_row": row_number,
                "work_zone": str(values[0]),
                "road_name": str(values[1] or ""),
                "subsection": str(values[2] or ""),
                "control_location": str(values[3] or ""),
                "control_start_date": parse_date(values[4]),
                "control_end_date": parse_date(values[5]),
                "control_hours": str(values[6] or ""),
                "control_method": str(values[7] or ""),
                "source_status": str(values[8] or ""),
                "source_note": str(values[9] or ""),
                "source_consistency_note": source_warnings.get(row_number, ""),
            }
        )
    if len(events) != EXPECTED_EVENTS:
        raise RuntimeError(f"Expected {EXPECTED_EVENTS} construction events, found {len(events)}")

    bus_sheet = workbook["버스전용차로 유예현황"]
    bus_rows: list[dict[str, Any]] = []
    for row_number in range(5, 14):
        values = [bus_sheet.cell(row_number, column).value for column in range(1, 8)]
        if not values[0]:
            continue
        end_text = str(values[4] or "")
        bus_rows.append(
            {
                "source_row": row_number,
                "road_name": str(values[0]),
                "relief_location": str(values[1] or ""),
                "length": str(values[2] or ""),
                "relief_start_date": parse_date(values[3]),
                "relief_end_or_status": end_text,
                "related_work_zone": str(values[5] or ""),
                "note": str(values[6] or ""),
            }
        )
    workbook.close()
    return events, bus_rows


def both(
    source_row: int,
    segments: Iterable[str],
    confidence: str,
    basis: str,
    impact_scope: str = "DIRECT",
    primary_analysis: int = 1,
) -> list[dict[str, Any]]:
    return [
        {
            "source_row": source_row,
            "segment_id": segment,
            "direction": direction,
            "impact_scope": impact_scope,
            "mapping_confidence": confidence,
            "mapping_basis": basis,
            "primary_analysis": primary_analysis,
        }
        for segment in segments
        for direction in ("AB", "BA")
    ]


def one(
    source_row: int,
    segment: str,
    direction: str,
    confidence: str,
    basis: str,
    impact_scope: str = "DIRECT",
    primary_analysis: int = 1,
) -> dict[str, Any]:
    return {
        "source_row": source_row,
        "segment_id": segment,
        "direction": direction,
        "impact_scope": impact_scope,
        "mapping_confidence": confidence,
        "mapping_basis": basis,
        "primary_analysis": primary_analysis,
    }


def build_mapping_rules() -> list[dict[str, Any]]:
    """Manual, auditable crosswalk from source control rows to 89 DTW units."""
    rules: list[dict[str, Any]] = []
    rules += [one(5, "SEG_44_243_244", "AB", "HIGH", "신탄진로 연축동 방향과 읍내→연축 진행방향 일치")]
    rules += both(6, ["SEG_43_242_243", "SEG_44_243_244"], "MEDIUM", "읍내삼거리 경계부 양쪽 인접 트램 구간")
    rules += both(7, ["SEG_43_242_243"], "HIGH", "계족로 동부여성가족원~읍내 보도육교")
    rules += [one(8, "SEG_44_243_244", "AB", "HIGH", "신탄진로 알뜰주유소~성우보육원, 신탄진IC 방향")]
    rules += both(9, ["SEG_42_241_242"], "MEDIUM", "계족로 보람아파트~법동네거리 구간")
    rules += [one(10, "SEG_42_241_242", "AB", "MEDIUM", "법동 방향 통제")]
    rules += [
        one(11, "SEG_12_212_213", "AB", "MEDIUM", "동부네거리→중부네거리 방향"),
        one(11, "SEG_13_213_214", "BA", "HIGH", "농수산오거리→중리네거리 방향"),
        one(11, "SEG_14_214_215", "BA", "HIGH", "농수산오거리→중리네거리 방향"),
    ]
    rules += both(12, ["SEG_13_213_214", "SEG_14_214_215"], "HIGH", "3공구 정거장 213~215 및 한밭대로 통제 위치")
    rules += both(13, ["SEG_16_216_217", "SEG_17_217_218", "SEG_18_218_219"], "HIGH", "선사유적네거리~정부청사~한밭대로네거리")
    rules += both(14, ["SEG_19_219_220", "SEG_20_220_221"], "HIGH", "선사유적네거리~대덕대교네거리")
    rules += both(15, ["SEG_21_221_222", "SEG_22_222_223", "SEG_23_223_224"], "MEDIUM", "국립중앙과학관~충대정문오거리")
    rules += both(16, ["SEG_22_222_223"], "HIGH", "구성삼거리~카이스트교삼거리")
    rules += both(17, ["SEG_26_226_227"], "LOW", "도안5단지네거리~상대지하차도 시점부; 원천 일정 표기 차이")
    rules += both(18, ["SEG_25_225_226"], "LOW", "유성네거리~도안5단지네거리; 원천 일정 표기 차이")
    rules += both(19, ["SEG_24_224_225"], "HIGH", "대학로 온천교 궁동~봉명동")
    rules += both(20, ["SEG_45_233_245"], "HIGH", "구봉중삼거리~진잠네거리")
    rules += [
        one(21, "SEG_36_236_237", "AB", "HIGH", "정림삼거리→도마삼거리 도마동 방향"),
        one(21, "SEG_37_237_238", "AB", "HIGH", "정림삼거리→도마삼거리 도마동 방향"),
    ]
    rules += both(22, ["SEG_36_236_237", "SEG_37_237_238"], "HIGH", "정림삼거리~도마삼거리 양방향")
    rules += both(23, ["SEG_01_201_202"], "HIGH", "서대전역네거리~서대전네거리")
    rules += both(24, ["SEG_02_202_203"], "MEDIUM", "홈플러스문화점~기독교연합봉사회관")
    rules += both(25, ["SEG_03_203_204"], "HIGH", "충무로 테미삼거리~보문산공원오거리")
    rules += [one(26, "SEG_07_207_208", "BA", "MEDIUM", "죽전네거리~신안네거리; AB는 DTW 제외 단위")]
    rules += both(27, ["SEG_04_204_205"], "MEDIUM", "보문교 하부 우회 영향 인접 구간", "ADJACENT", 0)
    rules += [one(28, "SEG_06_206_207", "BA", "HIGH", "죽전네거리→대전역네거리 방향")]
    rules += both(29, ["SEG_09_209_210"], "HIGH", "우송삼거리~가양네거리")
    return rules


def load_units() -> tuple[list[dict[str, str]], dict[tuple[str, str], dict[str, str]]]:
    with UNITS_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != EXPECTED_UNITS:
        raise RuntimeError(f"Expected {EXPECTED_UNITS} units, found {len(rows)}")
    lookup = {(row["segment_id"], row["direction"]): row for row in rows}
    if len(lookup) != EXPECTED_UNITS:
        raise RuntimeError("Duplicate spatial units")
    return rows, lookup


def join_mappings(
    events: list[dict[str, Any]],
    unit_lookup: dict[tuple[str, str], dict[str, str]],
) -> list[dict[str, Any]]:
    by_row = {event["source_row"]: event for event in events}
    rules = build_mapping_rules()
    mapped_rows: list[dict[str, Any]] = []
    for rule in rules:
        key = (rule["segment_id"], rule["direction"])
        if key not in unit_lookup:
            raise RuntimeError(f"Mapping targets an unavailable DTW unit: {key}")
        event = by_row[rule["source_row"]]
        unit = unit_lookup[key]
        mapped_rows.append(
            {
                "event_id": event["event_id"],
                "source_row": event["source_row"],
                "work_zone": event["work_zone"],
                "road_name": event["road_name"],
                "control_location": event["control_location"],
                "control_start_date": event["control_start_date"],
                "control_end_date": event["control_end_date"],
                "segment_id": rule["segment_id"],
                "direction": rule["direction"],
                "from_station_no": unit["from_station_no"],
                "from_station_name": unit["from_station_name"],
                "to_station_no": unit["to_station_no"],
                "to_station_name": unit["to_station_name"],
                "impact_scope": rule["impact_scope"],
                "mapping_confidence": rule["mapping_confidence"],
                "mapping_basis": rule["mapping_basis"],
                "primary_analysis": rule["primary_analysis"],
                "source_consistency_note": event["source_consistency_note"],
            }
        )
    if {event["source_row"] for event in events} - {row["source_row"] for row in mapped_rows}:
        raise RuntimeError("At least one construction event lacks a mapping rule")
    return mapped_rows


def load_labeled_daily_samples() -> tuple[list[dict[str, Any]], dict[tuple[str, str], list[dict[str, Any]]], list[str]]:
    labels: dict[str, dict[str, str]] = {}
    with LABELS_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            labels[row["sample_id"]] = row
    if len(labels) != EXPECTED_SAMPLES:
        raise RuntimeError("Unexpected final-label row count")

    samples: list[dict[str, Any]] = []
    by_unit: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    with gzip.open(WINDOWS_PATH, "rt", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        speed_columns = [field for field in (reader.fieldnames or []) if field.startswith("speed_index_")]
        if len(speed_columns) != EXPECTED_BINS:
            raise RuntimeError("Daily-window file no longer has 96 speed bins")
        for row in reader:
            label = labels.pop(row["sample_id"], None)
            if label is None:
                raise RuntimeError(f"Missing cluster label for {row['sample_id']}")
            values = [float(row[column]) for column in speed_columns]
            sample = {
                "sample_id": row["sample_id"],
                "service_date": parse_date(row["service_date"]),
                "year_month": row["year_month"],
                "segment_id": row["segment_id"],
                "direction": row["direction"],
                "from_station_no": row["from_station_no"],
                "from_station_name": row["from_station_name"],
                "to_station_no": row["to_station_no"],
                "to_station_name": row["to_station_name"],
                "cluster_id": int(label["cluster_id"]),
                "cluster_label": label["cluster_label"],
                "mean_speed_index": sum(values) / EXPECTED_BINS,
                "low_speed": int(label["cluster_label"] in LOW_SPEED_LABELS),
                "severe": int(label["cluster_label"] == SEVERE_LABEL),
            }
            samples.append(sample)
            by_unit[(sample["segment_id"], sample["direction"])].append(sample)
    if labels or len(samples) != EXPECTED_SAMPLES:
        raise RuntimeError("Daily windows and final labels do not form a one-to-one join")
    for unit_samples in by_unit.values():
        unit_samples.sort(key=lambda row: row["service_date"])
    if len(by_unit) != EXPECTED_UNITS:
        raise RuntimeError("Not all 89 units are represented in daily samples")
    return samples, by_unit, speed_columns


def summarize_samples(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {
            "sample_count": 0,
            "mean_speed_index": None,
            "low_speed_share": None,
            "severe_share": None,
            "dominant_cluster_label": "",
        }
    counts = Counter(row["cluster_label"] for row in rows)
    dominant = sorted(counts, key=lambda label: (-counts[label], label))[0]
    return {
        "sample_count": len(rows),
        "mean_speed_index": statistics.fmean(row["mean_speed_index"] for row in rows),
        "low_speed_share": statistics.fmean(row["low_speed"] for row in rows),
        "severe_share": statistics.fmean(row["severe"] for row in rows),
        "dominant_cluster_label": dominant,
    }


def create_monthly_summaries(
    samples: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_unit_month: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    by_month: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sample in samples:
        by_unit_month[(sample["year_month"], sample["segment_id"], sample["direction"])].append(sample)
        by_month[sample["year_month"]].append(sample)

    labels = ["시간대 집중 정체형", "전일 원활형", "출퇴근 집중형", "상시 저속형"]
    monthly_rows: list[dict[str, Any]] = []
    for (year_month, segment_id, direction), rows in sorted(by_unit_month.items()):
        counts = Counter(row["cluster_label"] for row in rows)
        dominant = sorted(labels, key=lambda label: (-counts[label], label))[0]
        first = rows[0]
        monthly_rows.append(
            {
                "year_month": year_month,
                "segment_id": segment_id,
                "direction": direction,
                "from_station_no": first["from_station_no"],
                "from_station_name": first["from_station_name"],
                "to_station_no": first["to_station_no"],
                "to_station_name": first["to_station_name"],
                "complete_day_count": len(rows),
                "mean_daily_speed_index": statistics.fmean(row["mean_speed_index"] for row in rows),
                "time_specific_count": counts["시간대 집중 정체형"],
                "time_specific_share": counts["시간대 집중 정체형"] / len(rows),
                "all_day_smooth_count": counts["전일 원활형"],
                "all_day_smooth_share": counts["전일 원활형"] / len(rows),
                "commute_peak_count": counts["출퇴근 집중형"],
                "commute_peak_share": counts["출퇴근 집중형"] / len(rows),
                "persistent_low_count": counts["상시 저속형"],
                "persistent_low_share": counts["상시 저속형"] / len(rows),
                "low_speed_cluster_count": counts["출퇴근 집중형"] + counts["상시 저속형"],
                "low_speed_cluster_share": (counts["출퇴근 집중형"] + counts["상시 저속형"]) / len(rows),
                "dominant_cluster_label": dominant,
            }
        )

    network_rows: list[dict[str, Any]] = []
    for year_month, rows in sorted(by_month.items()):
        counts = Counter(row["cluster_label"] for row in rows)
        network_rows.append(
            {
                "year_month": year_month,
                "complete_window_count": len(rows),
                "represented_spatial_units": len({(row["segment_id"], row["direction"]) for row in rows}),
                "mean_daily_speed_index": statistics.fmean(row["mean_speed_index"] for row in rows),
                "time_specific_share": counts["시간대 집중 정체형"] / len(rows),
                "all_day_smooth_share": counts["전일 원활형"] / len(rows),
                "commute_peak_share": counts["출퇴근 집중형"] / len(rows),
                "persistent_low_share": counts["상시 저속형"] / len(rows),
                "low_speed_cluster_share": (counts["출퇴근 집중형"] + counts["상시 저속형"]) / len(rows),
                "is_partial_month": int(year_month == max(by_month)),
            }
        )
    return monthly_rows, network_rows


def filter_period(
    rows: list[dict[str, Any]], start_date: date, end_date: date
) -> list[dict[str, Any]]:
    return [row for row in rows if start_date <= row["service_date"] <= end_date]


def analyze_event_shifts(
    mappings: list[dict[str, Any]],
    by_unit: dict[tuple[str, str], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    data_min = min(row["service_date"] for rows in by_unit.values() for row in rows)
    data_max = max(row["service_date"] for rows in by_unit.values() for row in rows)
    target_units_by_event: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for mapping in mappings:
        if mapping["primary_analysis"]:
            target_units_by_event[mapping["event_id"]].add((mapping["segment_id"], mapping["direction"]))

    output: list[dict[str, Any]] = []
    for mapping in mappings:
        event_start = mapping["control_start_date"]
        event_end = mapping["control_end_date"]
        post_last = min(event_end, data_max, event_start + timedelta(days=MAX_WINDOW_DAYS - 1))
        available_post_days = max(0, (post_last - event_start).days + 1)
        available_pre_days = max(0, (event_start - data_min).days)
        window_days = min(MAX_WINDOW_DAYS, available_post_days, available_pre_days)
        if window_days > 0:
            pre_start = event_start - timedelta(days=window_days)
            pre_end = event_start - timedelta(days=1)
            post_start = event_start
            post_end = event_start + timedelta(days=window_days - 1)
        else:
            pre_start = pre_end = post_start = post_end = None

        unit_key = (mapping["segment_id"], mapping["direction"])
        pre_rows = filter_period(by_unit[unit_key], pre_start, pre_end) if pre_start else []
        post_rows = filter_period(by_unit[unit_key], post_start, post_end) if post_start else []
        before = summarize_samples(pre_rows)
        after = summarize_samples(post_rows)
        sufficient = int(
            mapping["primary_analysis"] == 1
            and before["sample_count"] >= MIN_PERIOD_SAMPLES
            and after["sample_count"] >= MIN_PERIOD_SAMPLES
        )

        control_speed_deltas: list[float] = []
        control_low_deltas: list[float] = []
        if sufficient:
            for control_key, control_samples in by_unit.items():
                if control_key in target_units_by_event[mapping["event_id"]]:
                    continue
                control_pre = summarize_samples(filter_period(control_samples, pre_start, pre_end))
                control_post = summarize_samples(filter_period(control_samples, post_start, post_end))
                if (
                    control_pre["sample_count"] >= MIN_PERIOD_SAMPLES
                    and control_post["sample_count"] >= MIN_PERIOD_SAMPLES
                ):
                    control_speed_deltas.append(
                        control_post["mean_speed_index"] - control_pre["mean_speed_index"]
                    )
                    control_low_deltas.append(
                        control_post["low_speed_share"] - control_pre["low_speed_share"]
                    )

        if sufficient and control_speed_deltas:
            speed_delta = after["mean_speed_index"] - before["mean_speed_index"]
            low_delta = after["low_speed_share"] - before["low_speed_share"]
            severe_delta = after["severe_share"] - before["severe_share"]
            control_speed_delta = statistics.median(control_speed_deltas)
            control_low_delta = statistics.median(control_low_deltas)
            did_speed_delta = speed_delta - control_speed_delta
            did_low_delta = low_delta - control_low_delta
            raw_flag = int(
                low_delta >= RAW_LOW_SHARE_DELTA_THRESHOLD
                and speed_delta <= RAW_SPEED_DELTA_THRESHOLD
            )
            rapid_flag = int(
                raw_flag
                and did_low_delta >= DID_LOW_SHARE_DELTA_THRESHOLD
                and did_speed_delta <= DID_SPEED_DELTA_THRESHOLD
            )
            if rapid_flag:
                status = "공사 후 급격 악화형"
            elif raw_flag:
                status = "원시 악화·공통 추세 가능"
            elif low_delta > 0 or speed_delta < 0:
                status = "부분 악화"
            else:
                status = "변화 제한/개선"
        else:
            speed_delta = low_delta = severe_delta = None
            control_speed_delta = control_low_delta = None
            did_speed_delta = did_low_delta = None
            raw_flag = rapid_flag = 0
            status = "인접 통제 참고" if not mapping["primary_analysis"] else "비교 표본 부족"

        output.append(
            {
                **mapping,
                "data_start_date": data_min,
                "data_end_date": data_max,
                "comparison_window_days": window_days,
                "before_start_date": pre_start or "",
                "before_end_date": pre_end or "",
                "after_start_date": post_start or "",
                "after_end_date": post_end or "",
                "before_sample_count": before["sample_count"],
                "after_sample_count": after["sample_count"],
                "before_mean_speed_index": before["mean_speed_index"] if before["sample_count"] else "",
                "after_mean_speed_index": after["mean_speed_index"] if after["sample_count"] else "",
                "speed_index_delta": speed_delta if speed_delta is not None else "",
                "before_low_speed_share": before["low_speed_share"] if before["sample_count"] else "",
                "after_low_speed_share": after["low_speed_share"] if after["sample_count"] else "",
                "low_speed_share_delta": low_delta if low_delta is not None else "",
                "before_persistent_low_share": before["severe_share"] if before["sample_count"] else "",
                "after_persistent_low_share": after["severe_share"] if after["sample_count"] else "",
                "persistent_low_share_delta": severe_delta if severe_delta is not None else "",
                "before_dominant_cluster": before["dominant_cluster_label"],
                "after_dominant_cluster": after["dominant_cluster_label"],
                "control_unit_count": len(control_speed_deltas),
                "control_median_speed_delta": control_speed_delta if control_speed_delta is not None else "",
                "control_median_low_share_delta": control_low_delta if control_low_delta is not None else "",
                "did_speed_index_delta": did_speed_delta if did_speed_delta is not None else "",
                "did_low_speed_share_delta": did_low_delta if did_low_delta is not None else "",
                "comparison_valid_flag": sufficient,
                "raw_deterioration_flag": raw_flag,
                "rapid_deterioration_flag": rapid_flag,
                "change_status": status,
            }
        )
    return output


def select_worst_event(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def score(row: dict[str, Any]) -> tuple[float, float, str]:
        if not row["comparison_valid_flag"]:
            return (-999.0, -999.0, row["event_id"])
        did_low = float(row["did_low_speed_share_delta"])
        did_speed = float(row["did_speed_index_delta"])
        return (
            100 * int(row["rapid_deterioration_flag"]) + did_low - did_speed,
            float(row["low_speed_share_delta"]) - float(row["speed_index_delta"]),
            row["event_id"],
        )
    return max(rows, key=score)


def aggregate_segment_shifts(
    units: list[dict[str, str]],
    event_shifts: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_unit: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    adjacent_by_unit: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in event_shifts:
        target = by_unit if row["primary_analysis"] else adjacent_by_unit
        target[(row["segment_id"], row["direction"])].append(row)

    output: list[dict[str, Any]] = []
    for unit in units:
        key = (unit["segment_id"], unit["direction"])
        direct = by_unit.get(key, [])
        adjacent = adjacent_by_unit.get(key, [])
        if direct:
            selected = select_worst_event(direct)
            mapping_status = "KNOWN_DIRECT_CONTROL"
            event_count = len({row["event_id"] for row in direct})
            valid_count = sum(int(row["comparison_valid_flag"]) for row in direct)
            rapid_count = sum(int(row["rapid_deterioration_flag"]) for row in direct)
            first_date = min(row["control_start_date"] for row in direct)
            last_date = max(row["control_start_date"] for row in direct)
            values = {
                "representative_event_id": selected["event_id"],
                "representative_work_zone": selected["work_zone"],
                "representative_control_location": selected["control_location"],
                "representative_control_start_date": selected["control_start_date"],
                "mapping_confidence": selected["mapping_confidence"],
                "comparison_window_days": selected["comparison_window_days"],
                "before_sample_count": selected["before_sample_count"],
                "after_sample_count": selected["after_sample_count"],
                "before_mean_speed_index": selected["before_mean_speed_index"],
                "after_mean_speed_index": selected["after_mean_speed_index"],
                "speed_index_delta": selected["speed_index_delta"],
                "before_low_speed_share": selected["before_low_speed_share"],
                "after_low_speed_share": selected["after_low_speed_share"],
                "low_speed_share_delta": selected["low_speed_share_delta"],
                "did_speed_index_delta": selected["did_speed_index_delta"],
                "did_low_speed_share_delta": selected["did_low_speed_share_delta"],
                "before_dominant_cluster": selected["before_dominant_cluster"],
                "after_dominant_cluster": selected["after_dominant_cluster"],
                "change_status": selected["change_status"],
                "rapid_deterioration_flag": int(rapid_count > 0),
                "source_consistency_note": selected["source_consistency_note"],
            }
        else:
            mapping_status = "ADJACENT_CONTROL_ONLY" if adjacent else "NO_KNOWN_DIRECT_CONTROL"
            event_count = valid_count = rapid_count = 0
            first_date = last_date = ""
            values = {
                "representative_event_id": adjacent[0]["event_id"] if adjacent else "",
                "representative_work_zone": adjacent[0]["work_zone"] if adjacent else "",
                "representative_control_location": adjacent[0]["control_location"] if adjacent else "",
                "representative_control_start_date": adjacent[0]["control_start_date"] if adjacent else "",
                "mapping_confidence": adjacent[0]["mapping_confidence"] if adjacent else "",
                "comparison_window_days": "",
                "before_sample_count": "",
                "after_sample_count": "",
                "before_mean_speed_index": "",
                "after_mean_speed_index": "",
                "speed_index_delta": "",
                "before_low_speed_share": "",
                "after_low_speed_share": "",
                "low_speed_share_delta": "",
                "did_speed_index_delta": "",
                "did_low_speed_share_delta": "",
                "before_dominant_cluster": "",
                "after_dominant_cluster": "",
                "change_status": "인접 통제 참고" if adjacent else "직접 통제 미매핑",
                "rapid_deterioration_flag": 0,
                "source_consistency_note": adjacent[0]["source_consistency_note"] if adjacent else "",
            }
        output.append(
            {
                "segment_id": unit["segment_id"],
                "direction": unit["direction"],
                "from_station_no": unit["from_station_no"],
                "from_station_name": unit["from_station_name"],
                "to_station_no": unit["to_station_no"],
                "to_station_name": unit["to_station_name"],
                "construction_mapping_status": mapping_status,
                "direct_control_event_count": event_count,
                "valid_comparison_event_count": valid_count,
                "rapid_deterioration_event_count": rapid_count,
                "first_control_start_date": first_date,
                "latest_control_start_date": last_date,
                **values,
            }
        )
    if len(output) != EXPECTED_UNITS:
        raise RuntimeError("Segment construction shift output does not contain 89 units")
    return output


def parameters_rows() -> list[dict[str, Any]]:
    return [
        {"parameter": "max_symmetric_window_days", "value": MAX_WINDOW_DAYS, "meaning": "착공일 전후 최대 대칭 비교 기간"},
        {"parameter": "minimum_samples_per_period", "value": MIN_PERIOD_SAMPLES, "meaning": "전·후 각각 필요한 완전 일별 윈도우 수"},
        {"parameter": "low_speed_clusters", "value": "출퇴근 집중형 + 상시 저속형", "meaning": "저속 군집 비중 정의"},
        {"parameter": "raw_low_share_delta_threshold", "value": RAW_LOW_SHARE_DELTA_THRESHOLD, "meaning": "원시 저속 군집 비중 증가 기준"},
        {"parameter": "raw_speed_delta_threshold", "value": RAW_SPEED_DELTA_THRESHOLD, "meaning": "원시 평균 속도지수 하락 기준"},
        {"parameter": "did_low_share_delta_threshold", "value": DID_LOW_SHARE_DELTA_THRESHOLD, "meaning": "공통 추세 보정 저속 비중 증가 기준"},
        {"parameter": "did_speed_delta_threshold", "value": DID_SPEED_DELTA_THRESHOLD, "meaning": "공통 추세 보정 속도지수 하락 기준"},
        {"parameter": "control_change", "value": "동일 기간 비대상 공간 단위 변화량 중앙값", "meaning": "네트워크 공통 추세의 보수적 대조값"},
        {"parameter": "rapid_deterioration_logic", "value": "네 가지 raw·DiD 기준 모두 충족", "meaning": "공사 후 급격 악화형 운영 판정"},
        {"parameter": "source_as_of", "value": "2026-07-08", "meaning": "공구별 통제현황 원천 기준일"},
    ]


def write_report(
    events: list[dict[str, Any]],
    mappings: list[dict[str, Any]],
    monthly_rows: list[dict[str, Any]],
    network_rows: list[dict[str, Any]],
    event_shifts: list[dict[str, Any]],
    segment_shifts: list[dict[str, Any]],
) -> None:
    direct_units = {
        (row["segment_id"], row["direction"])
        for row in mappings
        if row["primary_analysis"]
    }
    valid_events = [row for row in event_shifts if row["comparison_valid_flag"]]
    rapid_events = [row for row in event_shifts if row["rapid_deterioration_flag"]]
    rapid_units = [row for row in segment_shifts if row["rapid_deterioration_flag"]]
    status_counts = Counter(row["change_status"] for row in event_shifts if row["primary_analysis"])
    lines = [
        "# 공사 전후 군집 변화 분석 결과",
        "",
        f"- 공사 원천: `{CONSTRUCTION_XLSX.relative_to(ROOT)}` (2026-07-08 기준)",
        f"- 통제 이력: {len(events)}건, 직접 매핑 {sum(int(row['primary_analysis']) for row in mappings)}개 event×공간 단위",
        f"- 알려진 직접 통제 매핑 공간 단위: {len(direct_units)}/{EXPECTED_UNITS}",
        f"- 유효 전후 비교: {len(valid_events)}개 event×공간 단위",
        f"- 공사 후 급격 악화 event×공간 단위: {len(rapid_events)}개",
        f"- 공사 후 급격 악화 최종 공간 단위: {len(rapid_units)}개",
        f"- 월별 요약: {len(monthly_rows):,}행, {len(network_rows)}개월",
        "",
        "## 판정 규칙",
        "",
        f"- 통제 시작일 전후 최대 {MAX_WINDOW_DAYS}일을 대칭으로 비교하고, 통제 종료일과 데이터 종료일을 넘지 않음",
        f"- 전·후 완전 일별 윈도우가 각각 {MIN_PERIOD_SAMPLES}개 이상인 경우만 유효 비교",
        "- 저속 군집은 `출퇴근 집중형 + 상시 저속형`으로 정의",
        f"- 원시 변화: 저속 군집 비중 +{RAW_LOW_SHARE_DELTA_THRESHOLD:.0%}p 이상, 평균 속도지수 {RAW_SPEED_DELTA_THRESHOLD:.2f} 이하",
        f"- 공통 추세 보정 변화: 비대상 공간 단위 중앙값 대비 저속 비중 +{DID_LOW_SHARE_DELTA_THRESHOLD:.0%}p 이상, 속도지수 {DID_SPEED_DELTA_THRESHOLD:.2f} 이하",
        "- 네 조건을 모두 충족할 때만 `공사 후 급격 악화형`으로 판정",
        "",
        "## 직접 통제 이벤트 판정 분포",
        "",
    ]
    for status, count in sorted(status_counts.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"- {status}: {count}개")
    lines.extend(["", "## 급격 악화 공간 단위", ""])
    if rapid_units:
        for row in sorted(
            rapid_units,
            key=lambda value: (float(value["did_speed_index_delta"]), -float(value["did_low_speed_share_delta"])),
        ):
            lines.append(
                f"- {row['segment_id']} {row['direction']} "
                f"({row['from_station_name']}→{row['to_station_name']}): "
                f"저속 비중 {float(row['low_speed_share_delta']):+.1%}p, "
                f"속도지수 {float(row['speed_index_delta']):+.3f}, "
                f"DiD {float(row['did_low_speed_share_delta']):+.1%}p / {float(row['did_speed_index_delta']):+.3f}"
            )
    else:
        lines.append("- 보수적 raw·DiD 네 조건을 모두 충족한 공간 단위 없음")
    lines.extend(
        [
            "",
            "## 해석 한계",
            "",
            "- 이 결과는 통제 시작 전후의 관찰적 변화이며 공사 단독의 인과효과를 증명하지 않는다.",
            "- 8·11공구 통제 공지가 원천 워크북에 없어 해당 공구 구간은 직접 통제 판정 대상에서 빠질 수 있다.",
            "- 7공구 일부 단계의 위치·시작일은 현재 공식 공구 요약과 원천 워크북 표기가 달라 원천 값을 유지하되 LOW 신뢰도로 표시했다.",
            "- 13공구 동광장로 확장 구간은 실제 확장일이 미상이므로 최초 죽전~신안 구간만 매핑했다.",
            "- 버스전용차로 유예·폐지는 동시 교통정책으로 별도 파일에 보존했으며 공사 효과를 완화하거나 왜곡할 수 있다.",
            "- 2026-07 월별 값은 데이터 종료일인 2026-07-01 하루만 포함한 부분 월이다.",
            "",
            "## 검증",
            "",
            f"- 최종 라벨 {EXPECTED_SAMPLES:,}개와 일별 속도 벡터 {EXPECTED_SAMPLES:,}개의 1:1 결합 확인",
            f"- 공간 단위 {EXPECTED_UNITS}개 전부 최종 요약에 포함",
            f"- 통제 원천 {EXPECTED_EVENTS}건 모두 매핑 또는 인접 참고 규칙 보유",
            "- 전후 창은 대칭이며 통제 활성기간·데이터 기간을 벗어나지 않음",
            "- 월별 군집별 건수 합계가 월별 완전 윈도우 수와 일치",
            "",
            "## 산출물",
            "",
            "- `segment_month_cluster_summary.csv`: 월별 공간 단위 군집 비중과 평균 속도지수",
            "- `network_month_cluster_summary.csv`: 네트워크 전체 월별 군집 분포",
            "- `segment_construction_cluster_shift.csv`: 89개 공간 단위의 최종 공사 변화 판정",
            "- `segment_construction_event_shift.csv`: 통제 이벤트별 상세 전후·DiD 지표",
            "- `construction_control_events.csv`: 원천 통제 이력 정규화",
            "- `construction_segment_mapping.csv`: 통제 이벤트–공간 단위 매핑 근거",
            "- `bus_lane_relief_events.csv`: 동시 버스전용차로 유예·폐지 이력",
            "- `construction_analysis_parameters.csv`: 비교창과 판정 임계값",
            "- `construction_change_analysis.xlsx`: 핵심 KPI·급격 악화 구간·월별 추세·원천 매핑 통합문서",
        ]
    )
    (OUTPUT_DIR / "construction_change_qc_report.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8"
    )


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Loading construction source and 89-unit crosswalk", flush=True)
    events, bus_rows = load_source_events()
    units, unit_lookup = load_units()
    mappings = join_mappings(events, unit_lookup)
    print("Joining 36,154 final labels to daily speed-index windows", flush=True)
    samples, by_unit, _ = load_labeled_daily_samples()
    monthly_rows, network_rows = create_monthly_summaries(samples)
    print("Calculating symmetric before/after and control-adjusted changes", flush=True)
    event_shifts = analyze_event_shifts(mappings, by_unit)
    segment_shifts = aggregate_segment_shifts(units, event_shifts)

    write_csv(OUTPUT_DIR / "construction_control_events.csv", events)
    write_csv(OUTPUT_DIR / "bus_lane_relief_events.csv", bus_rows)
    write_csv(OUTPUT_DIR / "construction_segment_mapping.csv", mappings)
    write_csv(OUTPUT_DIR / "segment_month_cluster_summary.csv", monthly_rows)
    write_csv(OUTPUT_DIR / "network_month_cluster_summary.csv", network_rows)
    write_csv(OUTPUT_DIR / "segment_construction_event_shift.csv", event_shifts)
    write_csv(OUTPUT_DIR / "segment_construction_cluster_shift.csv", segment_shifts)
    write_csv(OUTPUT_DIR / "construction_analysis_parameters.csv", parameters_rows())
    write_report(events, mappings, monthly_rows, network_rows, event_shifts, segment_shifts)

    if sum(row["complete_window_count"] for row in network_rows) != EXPECTED_SAMPLES:
        raise RuntimeError("Network monthly totals do not reconcile to all labeled samples")
    for row in monthly_rows:
        cluster_total = (
            row["time_specific_count"] + row["all_day_smooth_count"]
            + row["commute_peak_count"] + row["persistent_low_count"]
        )
        if cluster_total != row["complete_day_count"]:
            raise RuntimeError("Monthly cluster counts do not reconcile")
    hashes = {
        path.name: sha256_file(path)
        for path in (
            OUTPUT_DIR / "segment_month_cluster_summary.csv",
            OUTPUT_DIR / "segment_construction_cluster_shift.csv",
            OUTPUT_DIR / "segment_construction_event_shift.csv",
        )
    }
    (OUTPUT_DIR / "output_sha256.txt").write_text(
        "\n".join(f"{digest}  {name}" for name, digest in hashes.items()) + "\n",
        encoding="utf-8",
    )
    rapid_count = sum(row["rapid_deterioration_flag"] for row in segment_shifts)
    valid_count = sum(row["comparison_valid_flag"] for row in event_shifts)
    print(
        f"Completed: events={len(events)}, mappings={len(mappings)}, valid={valid_count}, "
        f"rapid_units={rapid_count}",
        flush=True,
    )


if __name__ == "__main__":
    main()
